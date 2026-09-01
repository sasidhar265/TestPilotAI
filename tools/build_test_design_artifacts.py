"""Build the review-ready manual test workbook stored in test-artifacts/."""

# ruff: noqa: E501 -- Test descriptions remain readable as complete workbook cell values.

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT = Path(__file__).resolve().parents[1] / "test-artifacts" / "manual-test-cases.xlsx"

HEADERS = [
    "Test Case ID", "Title", "Objective", "Category", "Priority", "Execution Mode",
    "Feasibility Reason", "Preconditions", "Step No.", "Action", "Expected Result",
    "Test Data", "Acceptance Criteria", "Tags", "Status", "Actual Result", "Tester", "Evidence",
]

CASES = [
    {
        "id": "MTC-001", "title": "Generate manual cases from pasted requirements",
        "objective": "Confirm a tester can generate a complete, readable manual suite from valid text.",
        "category": "smoke", "priority": "P0",
        "reason": "Human review is required to assess usefulness, readability, and scenario relevance.",
        "pre": "Application is running; tester can access the web UI; use synthetic content only.",
        "data": "Requirement: A registered user can sign in and see their dashboard.",
        "ac": "FR-001, FR-004, FR-005", "tags": "manual, generation, ui",
        "steps": [
            ("Open the Quality Lifecycle Studio web UI.", "The requirement input and generation controls are visible."),
            ("Paste the synthetic requirement and select normal/manual output.", "The input is accepted and remains available for review."),
            ("Start generation and wait for completion.", "A suite of test cases is displayed without an unhandled error."),
            ("Inspect every generated case.", "Each case shows ID, title, objective, category, priority, execution mode, rationale, and observable steps."),
        ],
    },
    {
        "id": "MTC-002", "title": "Review manual and automation classification quality",
        "objective": "Verify classifications and feasibility reasons are specific to each scenario.",
        "category": "regression", "priority": "P1",
        "reason": "Classification quality depends on contextual human judgment.",
        "pre": "A mixed suite containing manual and automated cases has been generated.",
        "data": "Synthetic requirement containing deterministic API checks and visual usability checks.",
        "ac": "FR-005", "tags": "manual, classification, quality",
        "steps": [
            ("Review the execution mode on each generated case.", "Each case is explicitly marked manual or automation."),
            ("Compare each mode with the actions and expected results.", "Repeatable deterministic checks are automation candidates; subjective or human-dependent checks are manual."),
            ("Review every feasibility reason.", "Each reason explains the case-specific decision and is not generic boilerplate."),
        ],
    },
    {
        "id": "MTC-003", "title": "Upload an exported Apple Pages requirement",
        "objective": "Confirm actionable guidance and successful use of the supported export workflow.",
        "category": "sanity", "priority": "P1",
        "reason": "The workflow requires a person to export from Pages and judge the extraction preview.",
        "pre": "Apple Pages and the web UI are available; a synthetic Pages document exists.",
        "data": "Pages document with headings, bullets, and acceptance criteria; exported as PDF.",
        "ac": "FR-002", "tags": "manual, document, pages, pdf",
        "steps": [
            ("Attempt to select the native Pages document for upload.", "The UI or service provides actionable export guidance rather than silently processing it."),
            ("Export the same document from Pages as a text PDF.", "A readable PDF is created locally."),
            ("Upload the PDF and start generation.", "The document is accepted and a positive extracted-character count is reported."),
            ("Compare the generated suite with the source headings and criteria.", "The important source meaning is retained without invented sensitive data."),
        ],
    },
    {
        "id": "MTC-004", "title": "Assess OCR extraction from a photographed requirement",
        "objective": "Verify that a supported image produces usable requirement text or clear guidance.",
        "category": "regression", "priority": "P2",
        "reason": "A person must visually compare OCR output with the photographed source.",
        "pre": "OCR support is installed; image is within configured byte and pixel limits.",
        "data": "Synthetic JPEG containing printed acceptance criteria; no personal information.",
        "ac": "FR-002, FR-003", "tags": "manual, document, ocr, accessibility",
        "steps": [
            ("Upload the synthetic JPEG requirement.", "The filename and size are displayed before submission."),
            ("Start generation.", "The image is processed or an actionable OCR configuration message is shown."),
            ("Compare extracted meaning and generated coverage with the image.", "Text and scenarios preserve the legible criteria; any unreadable content is not falsely asserted."),
        ],
    },
    {
        "id": "MTC-005", "title": "Evaluate generated test quality and traceability",
        "objective": "Confirm the draft covers important criteria without duplicates or vague outcomes.",
        "category": "critical", "priority": "P0",
        "reason": "Final coverage and usefulness require accountable QA judgment.",
        "pre": "A suite and independent validation report have been generated.",
        "data": "Synthetic requirement with three uniquely labelled acceptance criteria.",
        "ac": "FR-007", "tags": "manual, validation, traceability",
        "steps": [
            ("Map every acceptance criterion to the cases claiming coverage.", "Each criterion has explicit, relevant coverage or a documented gap."),
            ("Compare case objectives and steps for semantic duplication.", "No two cases test the same condition and outcome without justification."),
            ("Review expected results for observability.", "Every result can be clearly judged pass or fail."),
            ("Compare observations with the validation report.", "The report identifies material coverage, duplicate, clarity, and expected-result issues."),
        ],
    },
    {
        "id": "MTC-006", "title": "Review the generated SpecFlow feature before handoff",
        "objective": "Confirm exported automation scenarios are understandable and implementation-ready.",
        "category": "smoke", "priority": "P1",
        "reason": "A human must assess domain language, maintainability, and suitability for the target framework.",
        "pre": "A BDD suite with automated cases has passed validation.",
        "data": "Synthetic login requirement with valid and invalid credential examples.",
        "ac": "FR-006, FR-011", "tags": "manual, bdd, specflow, review",
        "steps": [
            ("Download the suite as a .feature file.", "A feature artifact is downloaded with a meaningful Feature header."),
            ("Open the file in a Gherkin-aware editor.", "Feature, Scenario or Scenario Outline, Given, When, Then, and Examples syntax is recognized."),
            ("Review wording and data for the target application.", "Steps use consistent domain language, synthetic data, and observable outcomes."),
        ],
    },
    {
        "id": "MTC-007", "title": "Verify selective Jira publishing in the browser",
        "objective": "Confirm the user stays in control of exactly which cases are published.",
        "category": "critical", "priority": "P0",
        "reason": "External publishing requires explicit human approval and visual confirmation in Jira.",
        "pre": "Jira sandbox is configured; existing issue QA-101 is available; a validated suite is displayed.",
        "data": "Issue: QA-101; select MTC-001 and MTC-003 only.",
        "ac": "FR-012", "tags": "manual, jira, publishing, approval",
        "steps": [
            ("Clear all case selections, then select only the two specified cases.", "Only the intended checkboxes are selected."),
            ("Enter QA-101 and choose publish.", "A confirmation reports successful publication to QA-101."),
            ("Open QA-101 in the Jira sandbox and inspect attachments/comments.", "Only the selected cases are present; unselected cases were not published."),
        ],
    },
    {
        "id": "MTC-008", "title": "Check actionable and privacy-safe failure messages",
        "objective": "Verify common failures help the user recover without revealing confidential content.",
        "category": "critical", "priority": "P0",
        "reason": "Clarity and accidental disclosure require human inspection across the UI and logs.",
        "pre": "Tester can safely simulate authentication, quota, timeout, document, and Jira failures.",
        "data": "Synthetic secret marker: DO-NOT-DISPLAY-7F3; request ID: qa-safe-reference-101.",
        "ac": "FR-014, NFR-004, NFR-005", "tags": "manual, security, privacy, errors",
        "steps": [
            ("Trigger each supported failure using synthetic inputs.", "Each failure returns a concise category-specific recovery action and request ID."),
            ("Inspect the browser message, API response, and operational log entry.", "No requirement, prompt, generated content, credential, or synthetic secret marker is exposed."),
            ("Use the request ID to correlate the failure.", "The same safe identifier links the response and operational event."),
        ],
    },
    {
        "id": "MTC-009", "title": "Keyboard and screen-reader review of the generation workflow",
        "objective": "Assess the primary workflow against WCAG 2.2 AA interaction expectations.",
        "category": "regression", "priority": "P1",
        "reason": "Assistive-technology usability and focus behavior need human evaluation.",
        "pre": "Supported browser and screen reader are available; no mouse is used.",
        "data": "Synthetic valid requirement text.",
        "ac": "NFR-011", "tags": "manual, accessibility, wcag, ui",
        "steps": [
            ("Navigate the page using keyboard controls only.", "All interactive controls receive visible focus in a logical order."),
            ("Enter a requirement, choose options, and start generation.", "Labels, state changes, progress, and errors are announced meaningfully."),
            ("Review and select generated cases, then reach export controls.", "Expanded details and selection states are operable and understandable without a pointer."),
        ],
    },
    {
        "id": "MTC-010", "title": "Review agent documentation against runtime discovery",
        "objective": "Confirm users can understand every functional agent and its boundaries.",
        "category": "sanity", "priority": "P2",
        "reason": "Documentation completeness and comprehension are best assessed by a human reviewer.",
        "pre": "Application documentation and /api/agents are accessible.",
        "data": "No production data required.",
        "ac": "FR-013", "tags": "manual, documentation, agents",
        "steps": [
            ("Open the application documentation and retrieve /api/agents.", "Both sources list the functional agent pipeline."),
            ("Compare agent names, order, purpose, runtime, and capabilities.", "The documentation and endpoint agree, with no configured agent omitted."),
            ("Ask a new tester to explain when each agent runs.", "The tester can distinguish input, routing, generation, validation, conversion, output, and storage responsibilities."),
        ],
    },
]


def build() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Manual Test Cases"
    sheet.append(HEADERS)

    for case in CASES:
        for number, (action, expected) in enumerate(case["steps"], 1):
            sheet.append([
                case["id"], case["title"], case["objective"], case["category"], case["priority"],
                "manual", case["reason"], case["pre"], number, action, expected, case["data"],
                case["ac"], case["tags"], "Not Run", "", "", "",
            ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [14, 34, 40, 12, 10, 16, 45, 45, 10, 48, 48, 42, 22, 28, 12, 35, 18, 30]
    for column, width in zip(sheet.columns, widths, strict=True):
        sheet.column_dimensions[column[0].column_letter].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 32
    table = Table(displayName="ManualTestCases", ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)

    guide = workbook.create_sheet("Read Me")
    guide.append(["Manual test workbook", "Usage"])
    guide.append(["Source", "docs/company-solution-requirements.md"])
    guide.append(["Scope", "Human-executed scenarios; automated candidates are in automated-features/*.feature"])
    guide.append(["Data policy", "Synthetic data only. Never enter credentials, production personal data, or payment data."])
    guide.append(["Execution", "Set Status, Actual Result, Tester, and Evidence during an approved test run."])
    guide.append(["Disclaimer", "Test designs are drafts and are not evidence of execution."])
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 100
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    guide.freeze_panes = "A2"

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)


def verify() -> None:
    workbook = load_workbook(OUTPUT, read_only=True)
    sheet = workbook["Manual Test Cases"]
    assert sheet.max_row == 1 + sum(len(case["steps"]) for case in CASES)
    assert sheet.max_column == len(HEADERS)
    assert workbook.sheetnames == ["Manual Test Cases", "Read Me"]


if __name__ == "__main__":
    build()
    verify()
    print(f"Built and verified {OUTPUT}")
