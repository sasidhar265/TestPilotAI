"""Convert quality-gate-approved suites into Xray/Jira interchange files."""

import csv
import io
import json
import re
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment

from app.agents import AgentKind, FunctionalAgentDescriptor
from app.agents.test_case_validator import ValidationReport
from app.models import ExportFormat, TestSuite


def _short_step(value: str, limit: int = 100) -> str:
    first_clause = re.split(r"[.;\n]", " ".join(value.split()), maxsplit=1)[0].strip()
    if len(first_clause) <= limit:
        return first_clause
    return first_clause[: limit + 1].rsplit(" ", 1)[0].rstrip(",:") or first_clause[:limit]


class ContextConversionError(ValueError):
    """Raised when unapproved content reaches the conversion boundary."""


@dataclass(frozen=True)
class ConvertedArtifact:
    content: bytes
    filename: str
    media_type: str


class ContextConverterAgent:
    descriptor = FunctionalAgentDescriptor(
        id="context-converter-agent",
        name="Context Converter Agent",
        kind=AgentKind.CONTEXT_CONVERTER,
        purpose="Convert quality-gate-approved tests into Xray-ready CSV, Excel, or JSON.",
        runtime="local-deterministic",
        capabilities=(
            "xray-csv",
            "xray-excel",
            "xray-json",
            "gherkin-feature",
            "validated-input-only",
        ),
        instruction_file=".github/agents/context-converter.agent.md",
    )

    _HEADERS = (
        "Test Case Identifier",
        "Scenario Group",
        "Summary",
        "Test Type",
        "Priority",
        "Description",
        "Preconditions",
        "Step",
        "Action",
        "Test Data",
        "Expected Result",
        "Labels",
        "Requirements",
    )

    def convert(
        self, suite: TestSuite, validation: ValidationReport, output_format: ExportFormat
    ) -> ConvertedArtifact:
        if not validation.passed:
            raise ContextConversionError("Only a quality-gate-approved suite can be converted.")
        if output_format == ExportFormat.CSV:
            content = self._csv(suite)
            return ConvertedArtifact(content, "xray-test-cases.csv", "text/csv")
        if output_format == ExportFormat.EXCEL:
            content = self._excel(suite)
            return ConvertedArtifact(
                content,
                "xray-test-cases.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        if output_format == ExportFormat.FEATURE:
            content = self._feature(suite)
            return ConvertedArtifact(content, "automation-tests.feature", "text/x-gherkin")
        content = self._json(suite)
        return ConvertedArtifact(content, "xray-test-cases.json", "application/json")

    def _rows(self, suite: TestSuite) -> list[list[str | int]]:
        rows: list[list[str | int]] = []
        for case in suite.test_cases:
            data = "\n".join(f"{item.name}={item.value}" for item in case.test_data)
            for step_number, step in enumerate(case.steps, 1):
                rows.append(
                    [
                        case.id,
                        case.scenario_group,
                        case.title,
                        "Manual" if case.execution_mode.value == "manual" else "Generic",
                        case.priority,
                        case.objective,
                        "\n".join(case.preconditions),
                        step_number,
                        step.action,
                        data,
                        step.expected_result,
                        ",".join(case.tags),
                        ",".join(case.acceptance_criteria_covered),
                    ]
                )
        return rows

    def _csv(self, suite: TestSuite) -> bytes:
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        writer.writerow(self._HEADERS)
        writer.writerows(self._rows(suite))
        return output.getvalue().encode("utf-8-sig")

    def _excel(self, suite: TestSuite) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Xray Tests"
        sheet.append(self._HEADERS)
        for row in self._rows(suite):
            sheet.append(row)
        first_row = 2
        step_columns = {8, 9, 11}
        shared_columns = [
            column for column in range(1, len(self._HEADERS) + 1) if column not in step_columns
        ]
        for case in suite.test_cases:
            last_row = first_row + len(case.steps) - 1
            if case.execution_mode.value == "manual" and last_row > first_row:
                for column in shared_columns:
                    sheet.merge_cells(
                        start_row=first_row,
                        start_column=column,
                        end_row=last_row,
                        end_column=column,
                    )
            first_row = last_row + 1
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A2"
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _json(self, suite: TestSuite) -> bytes:
        tests = []
        for case in suite.test_cases:
            tests.append(
                {
                    "testtype": "Manual" if case.execution_mode.value == "manual" else "Generic",
                    "fields": {
                        "summary": case.title,
                        "description": case.objective,
                        "priority": {"name": case.priority},
                        "labels": case.tags,
                    },
                    "steps": [
                        {
                            "action": step.action,
                            "data": "\n".join(
                                f"{item.name}={item.value}" for item in case.test_data
                            ),
                            "result": step.expected_result,
                        }
                        for step in case.steps
                    ],
                    "requirements": case.acceptance_criteria_covered,
                }
            )
        return json.dumps(tests, indent=2, ensure_ascii=False).encode("utf-8")

    def _feature(self, suite: TestSuite) -> bytes:
        scenarios: list[str] = []
        for case in suite.test_cases:
            if case.execution_mode.value != "automation":
                continue
            scenario = (case.gherkin or "").strip()
            if not scenario.startswith(("Scenario:", "Scenario Outline:")):
                given = _short_step(
                    case.preconditions[0] if case.preconditions else "prerequisites are satisfied"
                )
                when = _short_step(case.steps[0].action)
                then = _short_step(case.steps[-1].expected_result)
                scenario = "\n".join(
                    [
                        f"Scenario: {case.title}",
                        f"  Given {given}",
                        f"  When {when}",
                        f"  Then {then}",
                    ]
                )
            if scenario.startswith("Scenario Outline:") and "Examples:" not in scenario:
                raise ContextConversionError(
                    f"Automation scenario outline {case.id} is missing an Examples table."
                )
            step_count = sum(
                line.strip().startswith(("Given ", "When ", "Then ", "And ", "But "))
                for line in scenario.splitlines()
            )
            if step_count > 4:
                raise ContextConversionError(
                    f"Automation scenario {case.id} exceeds the four-step Gherkin limit."
                )
            long_steps = [
                line.strip()
                for line in scenario.splitlines()
                if line.strip().startswith(("Given ", "When ", "Then ", "And ", "But "))
                and len(line.strip().split(maxsplit=1)[1]) > 100
            ]
            if long_steps:
                raise ContextConversionError(
                    f"Automation scenario {case.id} contains Gherkin step text over 100 characters."
                )
            scenarios.append(scenario)
        if not scenarios:
            raise ContextConversionError("The approved suite contains no automation scenarios.")
        background, scenarios = _extract_shared_background(scenarios)
        sections = [f"Feature: {suite.feature_name}"]
        if background:
            sections.append(f"Background:\n  {background}")
        sections.extend(scenarios)
        feature = "\n\n".join(sections) + "\n"
        return feature.encode("utf-8")


def _extract_shared_background(scenarios: list[str]) -> tuple[str | None, list[str]]:
    """Lift an identical leading Given into Background without changing scenario behavior."""
    if len(scenarios) < 2:
        return None, scenarios
    lines_by_scenario = [scenario.splitlines() for scenario in scenarios]
    given_indexes: list[int] = []
    given_steps: list[str] = []
    for lines in lines_by_scenario:
        index = next(
            (position for position, line in enumerate(lines) if line.strip().startswith("Given ")),
            -1,
        )
        if index < 0:
            return None, scenarios
        given_indexes.append(index)
        given_steps.append(lines[index].strip())
    if len(set(given_steps)) != 1:
        return None, scenarios
    reusable = []
    for lines, index in zip(lines_by_scenario, given_indexes, strict=True):
        reusable.append("\n".join(lines[:index] + lines[index + 1 :]))
    return given_steps[0], reusable
